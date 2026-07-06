# Аналитика по файлу titanic.xlsx

# ================================
# 1. СТЯГИВАНИЕ ДАННЫХ
# ================================
# 1.1 Стянуть данные из файла в питон, создать двухмерный массив с формой (кол-во пассажиров, кол-во столбцов-5)
# Например:
# Столбцы: id, age, gender, pclass, survived
# [
#     [1, 28, 0, 3, 1],
#     [2, 32, 1, 2, 1],
#     [3, 20, 0, 2, 0]
# ]
# 3 пассажира - 3 строки, 
# первый столбец это id - номер пассажира, 
# второй это возраст, 
# третий столбец 0-мужчина 1-женщина, 
# четвертый столбец класс пассажира - 1, 2 или 3
# пятый столбец это выжил пассажир или погиб, 1-выжил 0-погиб

from openpyxl import load_workbook, Workbook
import numpy as np
wb = load_workbook('titanic.xlsx')
ws = wb.active
# titanic = []
# for i in ws.iter_rows(values_only=True):
#     titanic.append(i)
# вариант Султана
ps = [list(i) for i in ws.iter_rows(values_only=True)] #[i,i,]
# print(ps)
# print(titanic)

ps.pop(0)
array_2D = np.array(ps)
# print(array_2D)

# array_2D = np.array(titanic[1:])
# print(array_2D)

# ================================
# 2. ФИЛЬТРАЦИЯ ДАННЫХ
# ================================

# 2.1
# Оставить только пассажиров, которые выжили (survived == 1)
survived = array_2D[array_2D[:,4]==1]
# print(survived)
# [array_2D[:,4] - вернет последний массив -> [1,1,0,0,1,...]
# [1,1,0,0,1,...] -> [True, True, False,...]
# array_2D[True, True, False,...] -> [[]]

# 2.2
# Найти всех женщин (gender == 1)
women = array_2D[array_2D[:,2]==1]
# print(women)

# 2.3
# Найти пассажиров 1 класса (pclass == 1)
pclass1 = array_2D[array_2D[:,3]==1]
# print(pclass1)
# 2.4
# Найти пассажиров старше 30 лет (age > 30)
older30 = array_2D[array_2D[:,1]>30]
# print(older30)
# 2.5
# Найти женщин из 1 класса, которые выжили
s_w_1 = array_2D[(array_2D[:,2]==1) & (array_2D[:,3]==1) & (array_2D[:,4]==1)]
# print(s_w_1[:,1].max()) # самая взрослая выжившая женщина в 1 классе
# print(s_w_1[s_w_1[:,1]]==(s_w_1[:,1].max()]) ?????? посмотреть правильный код
# ================================
# 3. АНАЛИТИКА ПО УСЛОВИЯМ
# ================================

# 3.1
# Сколько всего пассажиров в каждом классе (pclass)?
pclass1 = len(array_2D[(array_2D[:,3])==1])
pclass2 = len(array_2D[(array_2D[:,3])==2])
pclass3 = len(array_2D[(array_2D[:,3])==3])
# print(pclass1, pclass2, pclass3)
# 3.2
# Сколько мужчин и женщин в данных?
female = len(array_2D[(array_2D[:,2]==1)])
male = len(array_2D[(array_2D[:,2]==0)])
# print(female, male)

# 3.3
# Сколько выжило пассажиров в каждом классе?
cl1s = len(array_2D[(array_2D[:,3]==1) & (array_2D[:,4]==1)])
cl2s = len(array_2D[(array_2D[:,3]==2) & (array_2D[:,4]==1)])
cl3s = len(array_2D[(array_2D[:,3]==3) & (array_2D[:,4]==1)])
# print(cl1s, cl2s, cl3s)

# 3.4
# Средний возраст:
# - всех пассажиров
# - выживших
# - не выживших
total_avg_age = array_2D[:,1].mean()
# print(total_avg_age)
survived_avg = array_2D[array_2D[:,4] == 1][:,1].mean()
# print(survived_avg)
# ================================
# 4. ЛОГИЧЕСКИЕ УСЛОВИЯ (FILTERING)
# ================================

# 4.1
# Найти пассажиров:
# - женского пола И возраст < 18
# print(array_2D[(array_2D[:,2] ==1) & (array_2D[:,1] <18)])

# 4.2
# Найти пассажиров:
# - 1 класса ИЛИ возраст меньше 10
# print(array_2D[(array_2D[:,3] == 1) |(array_2D[:,1]<10)])
# 4.3
# Найти пассажиров, которые:
# - не выжили И были в 3 классе
print(array_2D[(array_2D[:,4] == 0) & (array_2D[:,3]==3)])
