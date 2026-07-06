import numpy as np

# list = [120000, 60000, 45000, 80000, 2000000]

# array_ = np.array([120000, 60000, 45000, 80000, 2000000]) # одномерный массив (строка)
# # print(array_)
# # print(array_.shape) # форма массива (кол-во)
# # print(array_.size) # размер массива (кол-во элем в массиве)

# # 2х мерный массив (строки, столбцы)
# array_2D = np.array([
#     [123000, 4500, 12000],
#     [200000, 5000, 80000],
#     [150000, 6000, 15000]
# ])
# # print(array_2D)
# # print(array_2D.shape) # форма массива (кол-во строк, кол-во столбцов)
# # print(array_2D.size) # размер массива (произвеение числе в форме, кол-во элем в массиве)

# # 3х мерный массив (таблицы, строки, столбцы)
# array_3D = np.array([
#     [
#         [1,3,4],
#         [4,1,9],
#         [6,7,2]
#     ],
#     [
#         [1,5,7],
#         [9,9,3],
#         [0,0,0]
#     ]
# ])
# # print(array_3D)
# # print(array_3D.shape) # форма массива (кол-во таблиц, кол-во строк, кол-во столбцов)
# # print(array_3D.size) # размер массива (произвеение числе в форме, кол-во элем в массиве)

# # при измененении размерности у массива, 
# # размер нового массива должен совпадать с размером старого массива 
# array_1D = np.array([28, 34, 22, 97, 45, 96]) #sh (6,), size ->6
# # print(array_1D)
# array_2D = array_1D.reshape((2,3)) #(3,2), (6,1), (1,6)
# # print(array_2D)
# array_3D = array_2D.reshape(3, 1, 2) #  (1, 2, 3), (2,1,3), (2,3,1)
# # print(array_3D)

# array_1D = np.array([28, 34, 22, 97, 45, 96, 34,12]) #sh (8,), size ->8
# # print(array_1D)
# array_3D = array_1D.reshape((2,2,2)) #sh (8,), size ->8
# # print(array_3D)

# array_3D = np.array([
#     [
#         [1,3,4],
#         [4,1,9],
#         [6,7,2]
#     ],
#     [
#         [1,5,7],
#         [9,9,3],
#         [0,0,0]
#     ]
# ])
# array_1D = array_3D.reshape((18,))
# # print(array_1D)
# array_2D = array_3D.reshape((2,9)) #(3,6), (6,3), (2,9) и т.д.
# print(array_2D)

# print(dir(array_1D))

# array_1D = np.array([2,34,1,232,1,343,5,10])
# # print(array_1D.max())  # сравнение по ряду

# array_2D = np.array([
#     [12,3,45],
#     [10,6,9]
# ])
# print(array_2D.max()) 
# print(array_2D.max(axis=0)) # сравнение вниз - движение вниз по столбцам #[ 12 6 45]
# print(array_2D.max(axis=1)) # движение вправо - сравнение строк [45 10]

# array_3D = np.array([
#     [
#         [12,0,1]
#         [10,23,0]
#     ],
#     [    [0,1,3],
#         [8,90,4]
#     ]    
    
# ])
# print(array_3D.max(axis=0)) # сравнение вниз - (строки между таблицами) движение вниз вглубь [[12 1 3] [10 90 4]]
# print(array_3D.max(axis=1)) # сравнение  - движение (строки внутри одной таблицы) [[12 23 1][8 90 4]]
# print(array_3D.max(axis=2)) #сравнение - движение в сторону (сравнение строк в одной таблице) [[12 23] [3 90]]

# 1. Стянуть данные с products листа и цены записать в отдельный лист

from openpyxl import load_workbook, Workbook


wb = load_workbook('products.xlsx')
ws = wb.active
prices = []
for row in ws.iter_rows(values_only=True):
    prices.append(row[1])

prices.remove('Цена')
# print(prices)

#2. Лист с ценами перевести в одномерный массив и найти макс, мин , среднюю цену продукта. 
# Также найти total price

array_1D = np.array(prices)
# print(array_1D)
# print(array_1D.max()) 
# print(array_1D.min())
# print(array_1D.mean())
# print(array_1D.sum())

#3. Перевести одномерный массив в двухмерный с любой формой, найти сумму между столбцами
# print(array_1D.shape)
# array_2D = array_1D.reshape((80,3))
# print(array_2D)

# print(array_2D.sum(axis=0))

######## Способы создания массива, Фильтрация  ########
#-----Способы создания массива----
# array_1D = np.array([12,34,65])  
# array_2D = np.zeros([4,5]) # создание пустых массивов
# array_3D = np.ones([4,5,5])
# # print(array_3D) 
# # точки [1. 1. 1. 1. 1.] на конце означают что числа float 
# array_1D = np.arange(5,12)
# print(array_1D)

#----Фильтрация----
# array_1D = np.array([12,34,65,12,4,3,4,231,2])
# # print(array_1D > 10) # выводит массив с булевыми значениями
# print((array_1D > 10) | (array_1D <20)) # & - это 'and' в numpy, | - 'or'
# print(array_1D[(array_1D > 20) & (array_1D%2==0)]) # выводит массив с цифровыми значениями

#4. Найти максимальный  и минимальный прайс у четных цен
# array_ch=array_1D[(array_1D%2==0)]
# print(array_ch)
# print(array_ch.max())
# print(array_ch.min())

# Аналитика по файлу titanic.xlsx
list = [123,123,12,53,123,5]
# print(list*2)
# print(list-5) # не позволяет 
array_1D = np.array([123,123,12,53,123,5])
# print(array_1D[2:5])
# print(array_1D[:])
# print(array_1D*2)
# print(array_1D-5)
array_2D = np.array([
    [12,4,5],
    [0,1,2],
    [23,2,4]
])
# print(array_2D[::2]) # поcледняя 2 - это шаг (взять все элементы с шагом 2)
# print(array_2D[:, 2]) # распечатать столбец под 2м индексом -> [5,2,4]
# print(array_2D *3) # -> [36,12,15],[0,3,6],[69,6,12]
res = array_2D[0]  #сохранили первую строку и сделали операцию сложения: print(res.sum()) 
# либо без сохранения -> print(array_2D[0].sum)
# print(res.sum())
print(array_2D[:,0].sum()) # сложит только элементы под нул индексом

